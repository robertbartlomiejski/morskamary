"""Tests for script CLI entry points and main functions."""

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

# Add scripts directory to path
sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))


class TestBuildTmbdDictionaryCLI:
    """Test build_tmbd_dictionary.py CLI functions."""

    def test_parse_args_default_sector(self):
        """parse_args should default to Blue Biotech sector."""
        import build_tmbd_dictionary

        with patch("sys.argv", ["build_tmbd_dictionary.py"]):
            args = build_tmbd_dictionary.parse_args()

        assert args.sector == "Blue Biotech"

    def test_parse_args_custom_sector(self):
        """parse_args should accept custom sector name."""
        import build_tmbd_dictionary

        with patch("sys.argv", ["build_tmbd_dictionary.py", "--sector", "Maritime Transport"]):
            args = build_tmbd_dictionary.parse_args()

        assert args.sector == "Maritime Transport"

    def test_parse_args_custom_output_dir(self):
        """parse_args should accept custom output directory."""
        import build_tmbd_dictionary
        from pathlib import Path

        with patch("sys.argv", ["build_tmbd_dictionary.py", "--output-dir", "/custom/path"]):
            args = build_tmbd_dictionary.parse_args()

        assert args.output_dir == Path("/custom/path")

    def test_load_literature_competence_extractor_success(self, tmp_path, monkeypatch):
        """load_literature_competence_extractor should load extraction function."""
        import build_tmbd_dictionary

        # Create a mock run_full_analysis.py module
        mock_module_path = tmp_path / "run_full_analysis.py"
        mock_module_path.write_text(
            "def extract_literature_competences():\n"
            "    return []"
        )

        monkeypatch.setattr(build_tmbd_dictionary, "REPO_ROOT", tmp_path)

        result = build_tmbd_dictionary.load_literature_competence_extractor()

        assert callable(result)

    def test_load_literature_competence_extractor_missing_file(self, tmp_path, monkeypatch):
        """load_literature_competence_extractor should raise ImportError if file missing."""
        import build_tmbd_dictionary

        # Create the file but it will fail on spec creation due to import issues
        # Instead, just verify the function checks for file existence
        monkeypatch.setattr(build_tmbd_dictionary, "REPO_ROOT", tmp_path)

        with pytest.raises((ImportError, FileNotFoundError)):
            build_tmbd_dictionary.load_literature_competence_extractor()

    def test_load_literature_competence_extractor_missing_function(self, tmp_path, monkeypatch):
        """load_literature_competence_extractor should raise ImportError if function missing."""
        import build_tmbd_dictionary

        # Create a mock run_full_analysis.py without the required function
        mock_module_path = tmp_path / "run_full_analysis.py"
        mock_module_path.write_text("# No extract_literature_competences function")

        monkeypatch.setattr(build_tmbd_dictionary, "REPO_ROOT", tmp_path)

        with pytest.raises(ImportError, match="must define a callable function"):
            build_tmbd_dictionary.load_literature_competence_extractor()

    def test_main_integration(self, tmp_path, monkeypatch, capsys):
        """main should orchestrate dictionary build and export."""
        import build_tmbd_dictionary

        # Mock the extract function
        def mock_extract():
            return []

        # Mock the output directory
        output_dir = tmp_path / "outputs"
        output_dir.mkdir()

        with patch("build_tmbd_dictionary.load_literature_competence_extractor") as mock_loader, \
             patch("build_tmbd_dictionary.LiteratureCompetenceRepository") as mock_repo_class, \
             patch("build_tmbd_dictionary.build_sector_dictionary_from_repository") as mock_build, \
             patch("build_tmbd_dictionary.export_sector_dictionary") as mock_export, \
             patch("sys.argv", ["build_tmbd_dictionary.py", "--sector", "Test Sector", "--output-dir", str(output_dir)]):

            mock_loader.return_value = mock_extract
            mock_repo = MagicMock()
            mock_repo_class.return_value = mock_repo
            mock_build.return_value = {"MARINE": [], "MARITIME": [], "OCEANIC": []}
            mock_export.return_value = output_dir / "test_sector_tmbd_dictionary.json"

            result = build_tmbd_dictionary.main()

            assert result == 0
            mock_loader.assert_called_once()
            mock_repo_class.assert_called_once_with(mock_extract)
            mock_build.assert_called_once_with(mock_repo, sector="Test Sector")
            mock_export.assert_called_once()

        captured = capsys.readouterr()
        assert "Test Sector" in captured.out
        assert "Competences:" in captured.out


class TestGenerateManifestCLI:
    """Test generate_manifest.py main function."""

    def test_main_generates_manifest(self, tmp_path, monkeypatch, capsys):
        """main should generate MANIFEST_SOURCES.csv."""
        import generate_manifest

        # Create test files
        (tmp_path / "test.py").touch()
        (tmp_path / "data").mkdir()
        (tmp_path / "data" / "test.csv").touch()

        manifest_path = tmp_path / "MANIFEST_SOURCES.csv"

        monkeypatch.setattr(generate_manifest, "REPO_ROOT", tmp_path)
        monkeypatch.setattr(generate_manifest, "MANIFEST_PATH", manifest_path)

        generate_manifest.main()

        assert manifest_path.exists()
        captured = capsys.readouterr()
        assert "Wrote" in captured.out
        assert str(manifest_path) in captured.out

    def test_main_preserves_manual_metadata(self, tmp_path, monkeypatch):
        """main should preserve manually added metadata from existing manifest."""
        import generate_manifest

        # Create test file
        test_file = tmp_path / "test.py"
        test_file.touch()

        manifest_path = tmp_path / "MANIFEST_SOURCES.csv"

        # Get the actual column headers from COLUMNS constant
        # Create existing manifest with manual metadata using correct columns
        manifest_path.write_text(
            "file_name,file_type,publisher_or_owner,year,version_or_identifier,"
            "licence_or_rights_note,summary_1_sentence,preferred_citation_key,"
            "text_available\n"
            "test.py,script,TestPublisher,2026,v1.0,MIT,Test summary,test_citation,yes\n"
        )

        monkeypatch.setattr(generate_manifest, "REPO_ROOT", tmp_path)
        monkeypatch.setattr(generate_manifest, "MANIFEST_PATH", manifest_path)

        generate_manifest.main()

        content = manifest_path.read_text()
        assert "2026" in content  # year preserved
        assert "TestPublisher" in content  # publisher preserved
        assert "test_citation" in content  # citation preserved

    def test_scan_files_skips_ignored_dirs(self, tmp_path, monkeypatch):
        """scan_files should skip .git, __pycache__, and other ignored directories."""
        import generate_manifest

        (tmp_path / ".git").mkdir()
        (tmp_path / ".git" / "config").touch()
        (tmp_path / "__pycache__").mkdir()
        (tmp_path / "__pycache__" / "test.pyc").touch()
        (tmp_path / "valid.py").touch()

        monkeypatch.setattr(generate_manifest, "REPO_ROOT", tmp_path)

        files = generate_manifest.scan_files()

        file_names = [f.name for f in files]
        assert "config" not in file_names
        assert "test.pyc" not in file_names
        assert "valid.py" in file_names

    def test_scan_files_sorts_results(self, tmp_path, monkeypatch):
        """scan_files should return sorted results."""
        import generate_manifest

        (tmp_path / "z.py").touch()
        (tmp_path / "a.py").touch()
        (tmp_path / "m.py").touch()

        monkeypatch.setattr(generate_manifest, "REPO_ROOT", tmp_path)

        files = generate_manifest.scan_files()

        names = [f.name for f in files]
        assert names == ["a.py", "m.py", "z.py"]


class TestBuildDerivedCLI:
    """Test build_derived.py main function."""

    def test_main_with_no_excel_files(self, tmp_path, monkeypatch, capsys):
        """main should handle case where no Excel files are found."""
        import build_derived

        monkeypatch.setattr(build_derived, "REPO_ROOT", tmp_path)
        monkeypatch.setattr(build_derived, "DERIVED_DIR", tmp_path / "derived")

        build_derived.main()

        captured = capsys.readouterr()
        assert "No Excel files found" in captured.out

    def test_main_processes_excel_files(self, tmp_path, monkeypatch, capsys):
        """main should process Excel files found in repository."""
        import build_derived

        # Create test Excel file
        excel_file = tmp_path / "test.xlsx"

        # Create a simple Excel file using openpyxl
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["A2"] = "Value1"
        ws["B2"] = "Value2"
        wb.save(str(excel_file))

        derived_dir = tmp_path / "derived"
        derived_dir.mkdir()

        monkeypatch.setattr(build_derived, "REPO_ROOT", tmp_path)
        monkeypatch.setattr(build_derived, "DERIVED_DIR", derived_dir)

        build_derived.main()

        captured = capsys.readouterr()
        assert "Processed" in captured.out or "sheets" in captured.out

        # Check that CSV was created
        csv_files = list(derived_dir.glob("*.csv"))
        assert len(csv_files) > 0
